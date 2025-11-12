"""
Excel file updater for Equipment Maintenance Log
Updates the network Excel file when maintenance is completed via Slack
Supports both .xls and .xlsx formats
"""

import os
from datetime import datetime
from typing import Optional, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# For .xls file support
try:
    import xlrd
    from xlutils.copy import copy as xlutils_copy
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False


def load_excel_config():
    """Load Excel file path from config."""
    try:
        import json
        with open("config.json", 'r') as f:
            config = json.load(f)
            return config.get("excel_file_path", r"\\insitu-serv2022\NetServ_2\PRODUCTION\Equipment Maintenance Log\SLACK Equipment Maintenance LOG.xls")
    except:
        return r"\\insitu-serv2022\NetServ_2\PRODUCTION\Equipment Maintenance Log\SLACK Equipment Maintenance LOG.xls"


def update_excel_xlsx(workbook, target_sheet, step_numbers_to_tick, date, user_name, frequency_key):
    """Update .xlsx file using openpyxl."""
    # Find the header row
    header_row = None
    step_cols = {}
    date_col = None
    notes_col = None
    
    # Look for header row with step numbers and "Date" column (search up to row 30)
    for row in range(1, min(30, target_sheet.max_row + 1)):
        for col in range(1, target_sheet.max_column + 1):
            cell_value = str(target_sheet.cell(row, col).value or "").strip()
            
            # Check for step columns (1, 2, 3, etc.)
            # Try to match the step number - be flexible with formatting
            for step_num in step_numbers_to_tick:
                # Check for exact match, or step/task prefix, or just the number as part of text
                if (cell_value == str(step_num) or 
                    cell_value.strip() == str(step_num) or
                    f"step {step_num}" in cell_value.lower() or 
                    f"task {step_num}" in cell_value.lower() or
                    (cell_value.isdigit() and int(cell_value) == step_num)):
                    step_cols[step_num] = col
            
            # Check for Date column
            if "date" in cell_value.lower() and date_col is None:
                date_col = col
            
            # Check for Notes column
            if "note" in cell_value.lower() and notes_col is None:
                notes_col = col
        
        # If we found the step columns we need, this is likely the header row
        if len(step_cols) == len(step_numbers_to_tick):
            header_row = row
            break
    
    # If we didn't find the expected steps, try alternative step numbers
    # (e.g., if Excel has "everyday" tasks that aren't in JSON)
    if not step_cols:
        # Try alternative step numbers for monthly (2,3 instead of 1,2)
        if frequency_key == "monthly" and step_numbers_to_tick == [2, 3]:
            # Already tried 2,3, try 1,2 as fallback
            step_numbers_to_tick = [1, 2]
            step_cols = {}
            for row in range(1, min(30, target_sheet.max_row + 1)):
                for col in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row, col).value or "").strip()
                    for step_num in step_numbers_to_tick:
                        if cell_value == str(step_num):
                            step_cols[step_num] = col
                    if "date" in cell_value.lower() and date_col is None:
                        date_col = col
                    if "note" in cell_value.lower() and notes_col is None:
                        notes_col = col
                if len(step_cols) == len(step_numbers_to_tick):
                    header_row = row
                    break
    
    if not step_cols:
        return {
            "success": False,
            "message": f"Could not find step columns {step_numbers_to_tick} in sheet: {target_sheet.title}. Header might be at a different row."
        }
    
    # Find the next empty row
    last_row = target_sheet.max_row
    if date_col:
        for row in range(header_row + 1, target_sheet.max_row + 1):
            if target_sheet.cell(row, date_col).value:
                last_row = row
    
    entry_row = last_row + 1
    
    # Format date for Excel (MM/DD/YYYY)
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        excel_date = date_obj.strftime("%m/%d/%Y")
    except:
        excel_date = date
    
    # Write date
    if date_col:
        target_sheet.cell(entry_row, date_col).value = excel_date
        target_sheet.cell(entry_row, date_col).alignment = Alignment(horizontal='center')
    
    # Write checkmarks in step columns
    for step_num in step_numbers_to_tick:
        if step_num in step_cols:
            target_sheet.cell(entry_row, step_cols[step_num]).value = "✓"
            target_sheet.cell(entry_row, step_cols[step_num]).alignment = Alignment(horizontal='center')
    
    # Add user credentials to Notes column
    if notes_col:
        notes_text = f"{user_name} - {excel_date}"
        existing_notes = str(target_sheet.cell(entry_row, notes_col).value or "").strip()
        if existing_notes:
            notes_text = f"{existing_notes}; {notes_text}"
        target_sheet.cell(entry_row, notes_col).value = notes_text
    
    return {
        "success": True,
        "entry_row": entry_row
    }


def update_excel_xls(rb, workbook, target_sheet, target_sheet_name, step_numbers_to_tick, date, user_name, frequency_key):
    """Update .xls file using xlrd/xlutils."""
    rb_sheet = rb.sheet_by_name(target_sheet_name)
    
    # Find the header row
    header_row = None
    step_cols = {}
    date_col = None
    notes_col = None
    
    # Look for header row (search up to row 30)
    for row in range(min(30, rb_sheet.nrows)):
        for col in range(rb_sheet.ncols):
            cell_value = str(rb_sheet.cell_value(row, col) or "").strip()
            
            # Check for step columns
            # Try to match the step number - be flexible with formatting
            for step_num in step_numbers_to_tick:
                # Check for exact match, or step/task prefix, or just the number as part of text
                if (cell_value == str(step_num) or 
                    cell_value.strip() == str(step_num) or
                    f"step {step_num}" in cell_value.lower() or 
                    f"task {step_num}" in cell_value.lower() or
                    (cell_value.isdigit() and int(cell_value) == step_num)):
                    step_cols[step_num] = col
            
            # Check for Date column
            if "date" in cell_value.lower() and date_col is None:
                date_col = col
            
            # Check for Notes column
            if "note" in cell_value.lower() and notes_col is None:
                notes_col = col
        
        if len(step_cols) == len(step_numbers_to_tick):
            header_row = row
            break
    
    # If we didn't find the expected steps, try alternative step numbers
    if not step_cols:
        # Try alternative step numbers for monthly (2,3 instead of 1,2)
        if frequency_key == "monthly" and step_numbers_to_tick == [2, 3]:
            step_numbers_to_tick = [1, 2]
            step_cols = {}
            for row in range(min(30, rb_sheet.nrows)):
                for col in range(rb_sheet.ncols):
                    cell_value = str(rb_sheet.cell_value(row, col) or "").strip()
                    for step_num in step_numbers_to_tick:
                        if cell_value == str(step_num):
                            step_cols[step_num] = col
                    if "date" in cell_value.lower() and date_col is None:
                        date_col = col
                    if "note" in cell_value.lower() and notes_col is None:
                        notes_col = col
                if len(step_cols) == len(step_numbers_to_tick):
                    header_row = row
                    break
    
    if not step_cols:
        return {
            "success": False,
            "message": f"Could not find step columns {step_numbers_to_tick} in sheet: {target_sheet_name}. Header might be at a different row."
        }
    
    # Find the next empty row
    last_row = rb_sheet.nrows
    if date_col:
        for row in range(header_row + 1, rb_sheet.nrows):
            if rb_sheet.cell_value(row, date_col):
                last_row = row
    
    entry_row = last_row
    
    # Format date for Excel (MM/DD/YYYY)
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        excel_date = date_obj.strftime("%m/%d/%Y")
    except:
        excel_date = date
    
    # Write date
    if date_col is not None:
        target_sheet.write(entry_row, date_col, excel_date)
    
    # Write checkmarks in step columns
    for step_num in step_numbers_to_tick:
        if step_num in step_cols:
            target_sheet.write(entry_row, step_cols[step_num], "✓")
    
    # Add user credentials to Notes column
    if notes_col is not None:
        notes_text = f"{user_name} - {excel_date}"
        existing_notes = str(rb_sheet.cell_value(entry_row, notes_col) or "").strip()
        if existing_notes:
            notes_text = f"{existing_notes}; {notes_text}"
        target_sheet.write(entry_row, notes_col, notes_text)
    
    return {
        "success": True,
        "entry_row": entry_row
    }


def update_excel_maintenance(
    equipment_name: str,
    serial_number: str,
    frequency: str,
    date: str,
    user_name: str,
    excel_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    Update Excel file with maintenance entry.
    
    Args:
        equipment_name: Name of the equipment
        serial_number: Serial number of the equipment
        frequency: Maintenance frequency (monthly, bi_annual, annual)
        date: Date in YYYY-MM-DD format
        user_name: Initials or name of the user who performed maintenance
        
    Returns:
        Dict with success status and message
    """
    try:
        # Get Excel path from config if not provided
        if excel_path is None:
            excel_path = load_excel_config()
        
        # Convert .xls to .xlsx path if needed, or try both
        xlsx_path = excel_path.replace('.xls', '.xlsx')
        
        # Try to open the file
        if not os.path.exists(excel_path) and not os.path.exists(xlsx_path):
            return {
                "success": False,
                "message": f"Excel file not found at: {excel_path}. Please check network connection and file path."
            }
        
        # Use xlsx if available, otherwise xls
        file_path = xlsx_path if os.path.exists(xlsx_path) else excel_path
        is_xls_format = file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx')
        
        # Get step numbers to tick based on equipment maintenance schedule
        step_numbers_to_tick = []
        try:
            import json
            with open("equipment_data.json", 'r') as f:
                equipment_list = json.load(f)
            
            equipment_info = None
            for eq in equipment_list:
                if (serial_number and str(eq.get("serial_number", "")).strip().lower() == serial_number.strip().lower()) or \
                   (equipment_name and str(eq.get("equipment_name", "")).strip().lower() == equipment_name.strip().lower()):
                    equipment_info = eq
                    break
            
            maintenance_schedule = equipment_info.get("maintenance_schedule", {}) if equipment_info else {}
            frequency_key = frequency.lower().replace("-", "_")
            
            if equipment_info:
                # The Excel file may have "Everyday" tasks (step 1) that aren't in JSON
                # We need to detect this from the Excel file structure
                # For now, we'll try to read the Excel to see what steps exist
                # But first, let's try a smarter approach: check if step 1 exists in Excel
                
                # Try to determine step offset by checking if there's an "everyday" frequency
                # or by reading the Excel file structure
                step_offset = 0
                
                # Check if Excel might have everyday tasks (we'll detect this when reading Excel)
                # For now, assume monthly starts at step 2 if we have monthly tasks
                if frequency_key == "monthly":
                    # Monthly might be steps 2,3 if everyday exists, or 1,2 if not
                    # We'll try both when searching
                    step_numbers_to_tick = [2, 3]  # Try 2,3 first (most common case)
                elif frequency_key == "bi_annual":
                    # Count: everyday(1) + monthly(2) = 3, so bi-annual starts at 4
                    monthly_count = len(maintenance_schedule.get("monthly", {}).get("tasks", []))
                    step_numbers_to_tick = [3 + monthly_count, 4 + monthly_count, 5 + monthly_count][:len(maintenance_schedule.get("bi_annual", {}).get("tasks", []))]
                elif frequency_key == "annual":
                    # Count all previous
                    monthly_count = len(maintenance_schedule.get("monthly", {}).get("tasks", []))
                    bi_annual_count = len(maintenance_schedule.get("bi_annual", {}).get("tasks", []))
                    start_step = 1 + monthly_count + bi_annual_count  # +1 for potential everyday
                    annual_tasks = maintenance_schedule.get("annual", {}).get("tasks", [])
                    step_numbers_to_tick = list(range(start_step, start_step + len(annual_tasks)))
                else:
                    # Default calculation
                    all_frequencies = ["monthly", "bi_annual", "annual"]
                    step_counter = 1
                    for freq in all_frequencies:
                        freq_tasks = maintenance_schedule.get(freq, {}).get("tasks", [])
                        if freq == frequency_key:
                            for i in range(len(freq_tasks)):
                                step_numbers_to_tick.append(step_counter)
                                step_counter += 1
                        else:
                            step_counter += len(freq_tasks)
            else:
                # Fallback logic - but these might be wrong if Excel has "everyday" tasks
                # We'll try both sets and see which one works
                if frequency_key == "monthly":
                    step_numbers_to_tick = [2, 3]  # Try 2,3 first (if everyday exists)
                elif frequency_key == "bi_annual":
                    step_numbers_to_tick = [4, 5, 6]  # Try 4,5,6 (if everyday and monthly exist)
                elif frequency_key == "annual":
                    step_numbers_to_tick = [7]  # Try 7 (if all previous exist)
        except Exception as e:
            # Fallback
            if frequency_key == "monthly":
                step_numbers_to_tick = [1, 2]
            elif frequency_key == "bi_annual":
                step_numbers_to_tick = [1, 2]
            elif frequency_key == "annual":
                step_numbers_to_tick = [3]
        
        if not step_numbers_to_tick:
            return {
                "success": False,
                "message": f"Could not determine which step columns to tick for frequency: {frequency}"
            }
        
        # Load workbook and find sheet
        if is_xls_format:
            if not XLS_SUPPORT:
                return {
                    "success": False,
                    "message": "xlrd and xlutils are required for .xls files. Install with: pip install xlrd==1.2.0 xlutils"
                }
            
            try:
                rb = xlrd.open_workbook(file_path, formatting_info=True)
                workbook = xlutils_copy(rb)
                
                # Find target sheet
                target_sheet_name = None
                for sheet_name in rb.sheet_names():
                    sheet = rb.sheet_by_name(sheet_name)
                    for row in range(min(20, sheet.nrows)):
                        for col in range(min(10, sheet.ncols)):
                            cell_value = str(sheet.cell_value(row, col) or "").strip()
                            if serial_number and serial_number.lower() in cell_value.lower():
                                target_sheet_name = sheet_name
                                break
                            if equipment_name and equipment_name.lower() in cell_value.lower():
                                target_sheet_name = sheet_name
                                break
                        if target_sheet_name:
                            break
                    if target_sheet_name:
                        break
                
                if not target_sheet_name:
                    return {
                        "success": False,
                        "message": f"Could not find sheet for equipment: {equipment_name} (S/N: {serial_number})"
                    }
                
                target_sheet = workbook.get_sheet(target_sheet_name)
                result = update_excel_xls(rb, workbook, target_sheet, target_sheet_name, step_numbers_to_tick, date, user_name, frequency_key)
                
                if result['success']:
                    workbook.save(file_path)
                    return {
                        "success": True,
                        "message": f"Updated Excel file: {target_sheet_name}, Row {result['entry_row'] + 1}"
                    }
                else:
                    return result
                    
            except PermissionError:
                return {
                    "success": False,
                    "message": "Permission denied. File may be open in Excel or locked by another user."
                }
            except Exception as e:
                return {
                    "success": False,
                    "message": f"Error updating .xls file: {str(e)}"
                }
        else:
            # Handle .xlsx files
            try:
                workbook = openpyxl.load_workbook(file_path)
                
                # Find target sheet
                target_sheet = None
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in range(1, min(20, sheet.max_row + 1)):
                        for col in range(1, min(10, sheet.max_column + 1)):
                            cell_value = str(sheet.cell(row, col).value or "").strip()
                            if serial_number and serial_number.lower() in cell_value.lower():
                                target_sheet = sheet
                                break
                            if equipment_name and equipment_name.lower() in cell_value.lower():
                                target_sheet = sheet
                                break
                        if target_sheet:
                            break
                    if target_sheet:
                        break
                
                if not target_sheet:
                    return {
                        "success": False,
                        "message": f"Could not find sheet for equipment: {equipment_name} (S/N: {serial_number})"
                    }
                
                result = update_excel_xlsx(workbook, target_sheet, step_numbers_to_tick, date, user_name, frequency_key)
                
                if result['success']:
                    workbook.save(file_path)
                    return {
                        "success": True,
                        "message": f"Updated Excel file: {target_sheet.title}, Row {result['entry_row']}"
                    }
                else:
                    return result
                    
            except PermissionError:
                return {
                    "success": False,
                    "message": "Permission denied. File may be open in Excel or locked by another user."
                }
            except Exception as e:
                return {
                    "success": False,
                    "message": f"Error updating Excel file: {str(e)}"
                }
    
    except Exception as e:
        return {
            "success": False,
            "message": f"Error updating Excel: {str(e)}"
        }
